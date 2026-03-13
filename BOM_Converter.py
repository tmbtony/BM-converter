"""
BOM Converter - Python EXE Edition
API Key는 코드에 넣지 않고 프로그램 실행 시 입력창에서 입력 & 저장

빌드:
  pip install pyinstaller pdfplumber python-docx openpyxl xlrd requests
  pyinstaller --onefile --windowed --name=BOM_Converter BOM_Converter.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading, json, re, os, sys, requests

try:    import pdfplumber
except: pdfplumber = None
try:    import openpyxl
except: openpyxl = None
try:    from docx import Document
except: Document = None

# ──────────────────────────────────────────────
CURRENT_VERSION = "1.0.3"
BASE_DIR  = os.path.dirname(sys.executable if getattr(sys,'frozen',False) else __file__)
KEY_FILE  = os.path.join(BASE_DIR, "api_key.txt")  # EXE 옆에 저장

def load_key():
    try:
        with open(KEY_FILE, encoding="utf-8") as f: return f.read().strip()
    except: return ""

def save_key(k):
    try:
        with open(KEY_FILE, "w", encoding="utf-8") as f: f.write(k.strip())
    except: pass
# ──────────────────────────────────────────────

SYSTEM_PROMPT = """You are an expert in steel piping materials (flanges and fittings).
Parse a customer BOM and extract ALL items into a structured JSON array.
Output ONLY a JSON array. No explanation, no markdown fences.

Each item must have these exact fields:
- no: IMPORTANT - if the customer BOM has an item number (e.g. 1, 2, 167, 245, A-1, P-101) for this line, you MUST use that exact number as-is. Do NOT reassign or renumber. Only assign sequential integer continuing from offset if there is truly no item number in the BOM.
- category: "FLANGE", "BW FITTING", or "FORGED FITTING"
- material:
  1. Space (not dash) between code and grade: A182-F316L -> A182 F316L
  2. Dual cert: keep A/SA prefix as-is
  3. Dual grade slash shorten: F316/316L -> F316/L, WP304/304L -> WP304/L
  4. For WP91 or F91: always append TYPE1 or TYPE2. Default to TYPE1 if not specified. e.g. "A234 WP91 TYPE1", "A182 F91 TYPE2"
  5. CLASS notation for F11, F22, WP11, WP22 grades: always use "CL." format with a period between CL and number. e.g. "A182 F11 CL.1", "A182 F22 CL.3", "A234 WP11 CL.1", "A234 WP22 CL.3". Never write CL1 or CL 1 or CLASS 1 — always "CL.1" or "CL.2" or "CL.3".
  6. SPECTACLE BLIND plate material notation:
     - Carbon Steel plates: A516-70, A516-65, A516-60 (keep hyphen, e.g. "A516-70")
     - Stainless Steel plates: A240-304, A240-304L, A240-304/L, A240-316, A240-316L, A240-316/L (keep hyphen)
     - Alloy Steel plates: A387-11, A387-22, A387-91 (keep hyphen, append CL if present e.g. "A387-11 CL.2")
  7. Examples: A105, A182 F316L, A/SA234 WPB, SA234 WP11 CL.1, A182 F11 CL.1, A516-70

- item (use ONLY these values):
  FLANGE: WNRF, WNRTJ, BLRF, BLRTJ, SORF, SORTJ, LAP JOINT, SWRF, SWRTJ,
          ORIFICE WNRF, ORIFICE WNRTJ, SPECTACLE BLRF, SPECTACLE BLRTJ
  BW FITTING: 90 ELL, 90 ELS, 45 ELL, 45 ELS, TEE, BARRED TEE, CR, ER, CAP, STUB END
    For ELBOW/BEND radius: 1D or SR -> ELS/ELL(S), 1.5D or LR -> ELL, 3D/3R/5D/5R etc -> "[angle] [radius] BEND"
  FORGED FITTING: 90 ELBOW, 45 ELBOW, TEE, F-CPLG, H-CPLG, NIPPLE, CSN, ESN,
    WOL, SOL, TOL, NOL, LOL, EOL, FLANGEOLET, NIPOFLANGE, WELDOFLANGE,
    R-INSERT, HH-PLUG, RH-PLUG, SH-PLUG

- size1, size2:
  Fractions: 0.25->¼, 0.5->½, 0.75->¾, 1.25->1¼, 1.5->1½, 2.5->2½, 3.5->3½
  Text: 1/4->¼, 1/2->½, 3/4->¾, 1-1/4->1¼, 1-1/2->1½, 2-1/2->2½, 3-1/2->3½
  DN: DN25=1, DN32=1¼, DN40=1½, DN50=2, DN65=2½, DN80=3, DN100=4, DN125=5, DN150=6, DN200=8, DN250=10, DN300=12
  size2 for FLANGE: "" by default. Only fill size2 if it is a REDUCING FLANGE with two DIFFERENT sizes explicitly stated. If BOM shows same size twice (e.g. 24/24), leave size2 "".
  size2 for BW FITTING reducing items (ER, CR, reducing TEE, SPECTACLE BL): fill if different size present.
  OLET products: size1=larger(header), size2=smaller(branch). "MISSING" if size2 not provided.
  R-INSERT: size1=larger(pipe), size2=smaller(insert outlet). "MISSING" if size2 not provided.

- sch1:
  FLANGE: pressure rating as 150#, 300#, 600#, 900#, 1500#, 2500#. Convert CL/PN: PN10=150#, PN16=150#, PN25=300#, PN40=300#. "" if unknown.
  FORGED FITTING R-INSERT, HH-PLUG, RH-PLUG, SH-PLUG: pound rating as 3000#, 6000#, 9000#. "" if not provided (leave blank, do NOT set MISSING).
  FORGED FITTING others (except WOL): pressure rating as 3000#, 6000#, 9000#. "" if unknown.
  BW FITTING and WOL — CRITICAL SCHEDULE RULES:
    Standard names: STD, X-S, XX-S (always keep hyphen for X-S and XX-S)
    EXTRA STRONG / EXTRA HEAVY / XH / X-H / XS (when used as schedule name) -> X-S
    DOUBLE EXTRA STRONG / DOUBLE EXTRA HEAVY / XXH / XX-H / XXS (when used as schedule name) -> XX-S
    Numeric schedules — CRITICAL: carefully distinguish S40S vs S40, S80S vs S80, etc.:
      If BOM says "40S" or "SCH 40S" or "SCH40S" -> S40S (stainless schedule)
      If BOM says "40" or "SCH 40" or "SCH40" (no trailing S) -> S40
      If BOM says "80S" or "SCH 80S" -> S80S
      If BOM says "80" or "SCH 80" -> S80
      Other numeric schedules: S10, S10S, S20, S20S, S30, S60, S100, S120, S140, S160
      Always remove dashes from schedule numbers: S-40->S40, S-40S->S40S, S-80->S80
    mm thickness (no MWT keyword): format as xx.xx (e.g. 9.53, 12.70)
    CRITICAL inch thickness rule: ALWAYS append the inch symbol ("). Format as x.xxx" with exactly 3 decimal places minimum (e.g. 0.375", 0.500", 0.787", 1.000"). NEVER omit the " symbol. Check every inch thickness value before output.
    Thickness unit determination if not stated:
      Value < 4.0 -> INCH -> MUST append " (e.g. 2.265->2.265", 0.872->0.872", 1.622->1.622")
      Value >= 4.0 -> MM  -> no " (e.g. 9.53->9.53, 12.70->12.70)
    ONLY if BOM explicitly has "min wall","minimum wall","mwt","mw","minimum wall thickness": add MWT suffix.
      inch MWT: "x.xxx\" MWT" (e.g. "0.787\" MWT"). NEVER add MWT otherwise.
      mm MWT: "xx.xx MWT" (e.g. "9.53 MWT")
    WOL: MUST have schedule number. If BOM shows # rating for WELDOLET, IGNORE it. Set "MISSING" if no schedule found.
    "" if unknown.

- sch2:
  WNRF/WNRTJ/SWRF/SWRTJ/ORIFICE WNRF/ORIFICE WNRTJ: schedule/thickness (same rules as BW FITTING sch1 including inch " symbol). "MISSING" if not provided.
  BW FITTING reducing (CR/ER): secondary sch (same inch " rule). "" otherwise.
  Others: ""

- type:
  FLANGE: "" always
  BW FITTING: CRITICAL - this field must NEVER be left blank/empty. Always determine SMLS, WELDED, or WX:
    "SMLS" for seamless (keywords: SMLS, SEAMLESS, smls, seamless)
    "WELDED" for welded (keywords: WELDED, ERW, EFW, welded)
    "WX" for X-ray/RT 100% examined
    If manufacture method is truly not determinable from the BOM text -> set to "MISSING"
  FORGED FITTING:
    WOL -> "BW", SOL -> "SW", TOL -> "THRD"
    HH-PLUG, RH-PLUG, SH-PLUG -> "THRD" by default unless specified otherwise
    Others: "SW","THRD","PBE","TBE","POE X BOE","PLE X TSE". "" if unknown.

- etc:
  FLANGE size1>=26 (integer 26 or larger, e.g. 26,28,30,32...): "SERIES A" or "SERIES B" in etc. "MISSING" if not provided.
  FLANGE size1<26: etc="" (do NOT set MISSING for series)
  NIPPLE: length as "100mm L" or "3\" L". "MISSING" if not provided.
  BARRED TEE, CR, ER: "" always
  Add "GALV" if galvanizing/galvanized/HDG required.

- qty: integer
- For ALL fields: use "" for unknown/not applicable. NEVER use "-" as placeholder.

OLET exact name mapping (do NOT confuse):
  WELDOLET->WOL, SOCKETOLET->SOL, THREADOLET->TOL, NIPOLET->NOL,
  LATEROLET->LOL, ELBOLET->EOL, FLANGEOLET->FLANGEOLET,
  NIPOFLANGE->NIPOFLANGE, WELDOFLANGE->WELDOFLANGE (different from WELDOLET)
  REDUCING INSERT->R-INSERT, HEX HEAD PLUG->HH-PLUG, ROUND HEAD PLUG->RH-PLUG, SQUARE HEAD PLUG->SH-PLUG

Return [] if no items found."""

CHUNK_SIZE  = 60
COLS        = ["no","category","material","item","size1","size2","sch1","sch2","type","etc","qty"]
COL_HEADERS = ["No.","CATEGORY","MATERIAL","ITEM","Size 1 (Inch)","Size 2 (Inch)","Sch 1","Sch 2","TYPE","ETC","Qty.(EA)"]
COL_WIDTHS  = [40,100,110,140,80,80,80,80,90,130,70]

# ── 파일 파싱 ─────────────────────────────────
def parse_file(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".pdf":              return parse_pdf(path)
    elif ext in (".xlsx",".xls"): return parse_excel(path)
    elif ext == ".docx":          return parse_word(path)
    else:
        with open(path, encoding="utf-8", errors="ignore") as f: return f.read()

def parse_pdf(path):
    if not pdfplumber: return "[오류] pip install pdfplumber"
    text = ""
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            if tables:
                for tbl in tables:
                    for row in tbl: text += "\t".join([c or "" for c in row]) + "\n"
            else: text += (page.extract_text() or "") + "\n"
    return text

def parse_excel(path):
    lines = []
    if path.endswith(".xlsx") and openpyxl:
        wb = openpyxl.load_workbook(path, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                if any(c is not None for c in row):
                    lines.append("\t".join([str(c) if c is not None else "" for c in row]))
    else:
        try:
            import xlrd
            wb = xlrd.open_workbook(path)
            for ws in wb.sheets():
                for i in range(ws.nrows):
                    row = ws.row_values(i)
                    if any(str(c).strip() for c in row):
                        lines.append("\t".join([str(c) for c in row]))
        except Exception as e: return f"[xls 오류] {e}"
    return "\n".join(lines)

def parse_word(path):
    if not Document: return "[오류] pip install python-docx"
    doc = Document(path)
    lines = [p.text for p in doc.paragraphs if p.text.strip()]
    for tbl in doc.tables:
        for row in tbl.rows: lines.append("\t".join([c.text.strip() for c in row.cells]))
    return "\n".join(lines)

# ── 후처리 (클라이언트 사이드 안전망) ────────────
def post_process(rows):
    import re
    def fix_inch_thickness(val):
        if not val or not isinstance(val, str): return val
        val = val.strip()
        # 숫자.숫자 패턴 (MWT 포함 가능)
        m = re.match(r'^(\d+\.\d+)(\s*MWT)?$', val)
        if m:
            num = float(m.group(1))
            mwt = " MWT" if m.group(2) else ""
            # 3.0 미만 → 인치로 판단 → " 추가
            if num < 3.0:
                return f'{m.group(1)}"{mwt}'
        return val

    result = []
    for r in rows:
        row = dict(r)
        # BW FITTING type 빈값 → MISSING
        if row.get("category") == "BW FITTING":
            t = str(row.get("type", "")).strip()
            if not t or t == "-":
                row["type"] = "MISSING"
        # inch 두께 " 누락 보정
        row["sch1"] = fix_inch_thickness(row.get("sch1", ""))
        row["sch2"] = fix_inch_thickness(row.get("sch2", ""))
        result.append(row)
    return result

# ── API 호출 ──────────────────────────────────
def call_api(bom_chunk, offset, api_key):
    msgs = [{"role":"user","content":f"Starting item number: {offset+1}\n\nParse this BOM chunk:\n\n{bom_chunk}"}]
    full_text = ""
    while True:
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"Content-Type":"application/json","x-api-key":api_key,"anthropic-version":"2023-06-01"},
            json={"model":"claude-sonnet-4-20250514","max_tokens":8000,"system":SYSTEM_PROMPT,"messages":msgs},
            timeout=120,
        )
        data = resp.json()
        if "error" in data: raise Exception(data["error"]["message"])
        chunk = "".join(b.get("text","") for b in data.get("content",[]))
        full_text += chunk
        if data.get("stop_reason") != "max_tokens": break
        msgs.append({"role":"assistant","content":chunk})
        msgs.append({"role":"user","content":"Continue exactly from where you left off. Do not repeat any items."})
    text = re.sub(r"```json|```","",full_text).strip()
    if not text.endswith("]"):
        last = text.rfind("}"); text = (text[:last+1]+"]") if last!=-1 else "[]"
    if not text.startswith("["):
        s = text.find("["); text = text[s:] if s!=-1 else "[]"
    return json.loads(text)

# ── 메인 GUI ──────────────────────────────────
class BOMConverter(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"⚙  Piping BOM Converter  v{CURRENT_VERSION}")
        self.geometry("1200x860")
        self.resizable(True, True)
        self.configure(bg="#f0f4f8")
        self.rows = []
        self.abort = False
        self._build_ui()

    def _build_ui(self):
        # 헤더
        hdr = tk.Frame(self, bg="#1e3a5f", pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text="⚙  Piping BOM Converter", bg="#1e3a5f", fg="white",
                 font=("Segoe UI",16,"bold")).pack(side="left", padx=20)
        tk.Label(hdr, text=f"v{CURRENT_VERSION}  |  FLANGE / BW FITTING / FORGED FITTING",
                 bg="#1e3a5f", fg="#aac4e8", font=("Segoe UI",10)).pack(side="left")

        # API Key 입력
        api_frm = tk.Frame(self, bg="#fff8e1", pady=6)
        api_frm.pack(fill="x", padx=16, pady=(8,0))
        tk.Label(api_frm, text="🔑 API Key:", bg="#fff8e1",
                 font=("Segoe UI",9,"bold"), fg="#7c5c00").pack(side="left", padx=(8,4))
        self.api_var = tk.StringVar(value=load_key())
        api_entry = tk.Entry(api_frm, textvariable=self.api_var, width=70,
                             show="*", font=("Consolas",9), relief="solid", bd=1)
        api_entry.pack(side="left", padx=4)
        self._btn(api_frm, "👁 보기/숨기기",
                  lambda: api_entry.config(show="" if api_entry.cget("show")=="*" else "*"),
                  "#f0f0f0","#333").pack(side="left", padx=4)
        self._btn(api_frm, "💾 저장", lambda: [save_key(self.api_var.get()),
                  messagebox.showinfo("저장","API Key가 저장되었습니다!")],
                  "#e8f0fe","#1e3a5f").pack(side="left", padx=4)
        tk.Label(api_frm, text="※ 저장하면 다음 실행 시 자동 입력됩니다",
                 bg="#fff8e1", fg="#999", font=("Segoe UI",8)).pack(side="left", padx=8)

        # BOM 입력
        inp = tk.LabelFrame(self, text=" 📥  고객 BOM 입력 ", bg="#f0f4f8",
                            font=("Segoe UI",10,"bold"), fg="#1e3a5f", pady=8, padx=10)
        inp.pack(fill="x", padx=16, pady=(8,4))
        btn_row = tk.Frame(inp, bg="#f0f4f8")
        btn_row.pack(fill="x", pady=(0,6))
        self._btn(btn_row,"📎 파일 열기",self._open_file,"#e8f0fe","#1e3a5f").pack(side="left",padx=(0,6))
        self._btn(btn_row,"🗑 초기화",   self._clear,   "#fff0f0","#cc0000").pack(side="left",padx=(0,6))
        self.file_lbl = tk.Label(btn_row, text="", bg="#f0f4f8", fg="#666", font=("Segoe UI",9))
        self.file_lbl.pack(side="left")
        self.txt_input = scrolledtext.ScrolledText(inp, height=8, font=("Consolas",9),
                                                    wrap="none", relief="solid", bd=1)
        self.txt_input.pack(fill="x")

        # 컨트롤
        ctrl = tk.Frame(self, bg="#f0f4f8")
        ctrl.pack(fill="x", padx=16, pady=6)
        self.btn_convert = self._btn(ctrl,"🤖  AI 변환 시작",self._start_convert,"#1e3a5f","white",bold=True)
        self.btn_convert.pack(side="left")
        self.btn_stop = self._btn(ctrl,"⏹ 중단",self._stop,"#fff0f0","#cc0000")
        self.btn_stop.pack(side="left", padx=6)
        self.btn_stop.config(state="disabled")
        self._btn(ctrl,"📥 Excel 저장",self._save_excel,"#1e7a3f","white",bold=True).pack(side="right")
        self.status_var = tk.StringVar(value="BOM을 입력하고 변환 시작 버튼을 눌러주세요.")
        tk.Label(ctrl, textvariable=self.status_var, bg="#f0f4f8",
                 fg="#444", font=("Segoe UI",9)).pack(side="left", padx=12)

        self.progress = ttk.Progressbar(self, mode="determinate")
        self.progress.pack(fill="x", padx=16, pady=(0,4))

        # 결과 테이블
        res = tk.LabelFrame(self, text=" 📊  변환 결과 ", bg="#f0f4f8",
                            font=("Segoe UI",10,"bold"), fg="#1e3a5f", pady=6, padx=10)
        res.pack(fill="both", expand=True, padx=16, pady=(0,12))
        tf = tk.Frame(res, bg="#f0f4f8")
        tf.pack(fill="both", expand=True)
        self.tree = ttk.Treeview(tf, columns=COLS, show="headings", height=20)
        vsb = ttk.Scrollbar(tf, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(tf, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        for col, h, w in zip(COLS, COL_HEADERS, COL_WIDTHS):
            self.tree.heading(col, text=h)
            self.tree.column(col, width=w, anchor="center", minwidth=40)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tf.rowconfigure(0, weight=1); tf.columnconfigure(0, weight=1)
        self.tree.tag_configure("FLANGE",        background="#D6E4FF")
        self.tree.tag_configure("BW FITTING",     background="#FFF2CC")
        self.tree.tag_configure("FORGED FITTING", background="#D5F5E3")
        self.tree.tag_configure("MISSING_ROW",    background="#FFFDE7")
        style = ttk.Style(); style.theme_use("clam")
        style.configure("Treeview.Heading", background="#1e3a5f", foreground="white", font=("Segoe UI",9,"bold"))
        style.configure("Treeview", font=("Segoe UI",9), rowheight=24)

    def _btn(self, p, t, c, bg, fg, bold=False):
        return tk.Button(p, text=t, command=c, bg=bg, fg=fg, relief="flat",
                         padx=12, pady=6, cursor="hand2",
                         font=("Segoe UI",9,"bold" if bold else "normal"),
                         activebackground=bg, activeforeground=fg)

    def _open_file(self):
        path = filedialog.askopenfilename(
            filetypes=[("지원 파일","*.pdf *.xlsx *.xls *.docx *.txt *.csv"),("모든 파일","*.*")])
        if not path: return
        self.file_lbl.config(text=os.path.basename(path))
        self.status_var.set("파일 파싱 중..."); self.update()
        text = parse_file(path)
        self.txt_input.delete("1.0","end"); self.txt_input.insert("1.0", text)
        self.status_var.set(f"✅ 로드 완료: {os.path.basename(path)}")

    def _clear(self):
        self.txt_input.delete("1.0","end"); self.file_lbl.config(text="")
        self.rows.clear(); self.tree.delete(*self.tree.get_children())
        self.progress["value"] = 0; self.status_var.set("초기화 완료.")

    def _start_convert(self):
        bom = self.txt_input.get("1.0","end").strip()
        if not bom: messagebox.showwarning("입력 없음","BOM 내용을 입력해주세요."); return
        api_key = self.api_var.get().strip()
        if not api_key: messagebox.showwarning("API Key","API Key를 입력해주세요."); return
        self.abort = False; self.rows.clear(); self.tree.delete(*self.tree.get_children())
        self.btn_convert.config(state="disabled"); self.btn_stop.config(state="normal")
        threading.Thread(target=self._worker, args=(bom, api_key), daemon=True).start()

    def _stop(self): self.abort = True; self.status_var.set("⏹ 중단 요청됨...")

    def _worker(self, bom, api_key):
        lines = [l for l in bom.split("\n") if l.strip()]
        total = max(1,(len(lines)+CHUNK_SIZE-1)//CHUNK_SIZE)
        self.progress["maximum"] = total
        all_rows = []
        for i, start in enumerate(range(0,len(lines),CHUNK_SIZE)):
            if self.abort: break
            chunk = "\n".join(lines[start:start+CHUNK_SIZE])
            self.status_var.set(f"⏳ 청크 {i+1}/{total} 처리 중... (누적 {len(all_rows)}개)")
            self.progress["value"] = i+1; self.update_idletasks()
            try:
                parsed = call_api(chunk, len(all_rows), api_key)
                parsed = post_process(parsed)
                all_rows.extend(parsed)
                self.after(0, self._append_rows, parsed)
            except Exception as e:
                self.after(0, messagebox.showerror, "API 오류", f"청크 {i+1} 오류:\n{e}"); break
        self.rows = all_rows
        self.after(0, self._on_done)

    def _append_rows(self, new_rows):
        for r in new_rows:
            cat  = r.get("category","")
            vals = [r.get(c,"") for c in COLS]
            tag  = "MISSING_ROW" if any(str(v)=="MISSING" for v in vals) else cat
            self.tree.insert("","end", values=vals, tags=(tag,))

    def _on_done(self):
        self.btn_convert.config(state="normal"); self.btn_stop.config(state="disabled")
        cats = {c:sum(1 for r in self.rows if r.get("category")==c)
                for c in ["FLANGE","BW FITTING","FORGED FITTING"]}
        self.status_var.set(f"✅ 완료! 총 {len(self.rows)}개  |  FLANGE:{cats['FLANGE']}  BW:{cats['BW FITTING']}  FORGED:{cats['FORGED FITTING']}")
        self.progress["value"] = self.progress["maximum"]

    def _save_excel(self):
        if not self.rows: messagebox.showwarning("없음","변환된 데이터가 없습니다."); return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
               filetypes=[("Excel","*.xlsx")], initialfile="견적서.xlsx")
        if not path: return
        try:
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "견적서"
            thin = Side(style="thin",color="CCCCCC")
            border = Border(left=thin,right=thin,top=thin,bottom=thin)
            hdr_fill = PatternFill("solid",fgColor="1E3A5F")
            hdr_font = Font(bold=True,color="FFFFFF",size=10)
            for ci,h in enumerate(COL_HEADERS,1):
                c = ws.cell(row=1,column=ci,value=h)
                c.fill,c.font = hdr_fill,hdr_font
                c.alignment = Alignment(horizontal="center",vertical="center")
                c.border = border
            fill_map = {"FLANGE":PatternFill("solid",fgColor="D6E4FF"),
                        "BW FITTING":PatternFill("solid",fgColor="FFF2CC"),
                        "FORGED FITTING":PatternFill("solid",fgColor="D5F5E3")}
            mf = PatternFill("solid",fgColor="FFF176")
            mfont = Font(bold=True,color="B45309",size=10)
            for ri,row in enumerate(self.rows,2):
                cat = row.get("category","")
                rf  = fill_map.get(cat,PatternFill("solid",fgColor="F9FAFB"))
                for ci,col in enumerate(COLS,1):
                    val = row.get(col,"")
                    c   = ws.cell(row=ri,column=ci,value=val)
                    c.alignment = Alignment(horizontal="center",vertical="center")
                    c.border = border
                    c.fill   = mf if str(val)=="MISSING" else rf
                    if str(val)=="MISSING": c.font = mfont
            for ci,w in enumerate(COL_WIDTHS,1):
                ws.column_dimensions[get_column_letter(ci)].width = w/7
            ws.freeze_panes = "A2"
            wb.save(path)
            messagebox.showinfo("완료",f"저장 완료:\n{path}")
        except Exception as e: messagebox.showerror("오류",str(e))

if __name__ == "__main__":
    app = BOMConverter()
    app.mainloop()
